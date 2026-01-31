"""
PersGest Database Manager
Gestisce Excel come database con 18 tabelle (fogli)
"""

import pandas as pd
import threading
import contextlib
from datetime import datetime
import json
import time
import re
import os
import shutil
from pathlib import Path
from datetime import datetime


# --- Concorrenza / sicurezza scritture (Excel come DB) ---
_PERSGEST_WRITE_LOCK = threading.RLock()

@contextlib.contextmanager
def _persgest_write_lock(db_path: Path, timeout: int = 120, poll: float = 0.1):
    """Lock globale + lock su file (filesystem locale) per serializzare le scritture.

    Implementazione senza dipendenze esterne:
    - Windows: msvcrt.locking
    - Linux/macOS: fcntl.flock
    """
    with _PERSGEST_WRITE_LOCK:
        lock_file = db_path.with_suffix(db_path.suffix + ".lock")
        lock_file.parent.mkdir(parents=True, exist_ok=True)

        f = open(lock_file, "a+b")
        start = time.time()
        locked = False
        try:
            while True:
                try:
                    if os.name == "nt":
                        import msvcrt
                        msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                    else:
                        import fcntl
                        fcntl.flock(f, fcntl.LOCK_EX | fcntl.LOCK_NB)
                    locked = True
                    break
                except OSError:
                    if (time.time() - start) >= timeout:
                        raise TimeoutError(f"Timeout lock file: {lock_file}")
                    time.sleep(poll)

            yield
        finally:
            if locked:
                try:
                    if os.name == "nt":
                        import msvcrt
                        msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
                    else:
                        import fcntl
                        fcntl.flock(f, fcntl.LOCK_UN)
                except Exception:
                    pass
            try:
                f.close()
            except Exception:
                pass

def _backup_excel(db_path: Path, keep_last: int = 200):
    """Backup best-effort del DB prima di una scrittura; retention sugli ultimi N backup."""
    try:
        if not db_path.exists():
            return
        backups_dir = db_path.parent / "backups"
        backups_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_path = backups_dir / f"{db_path.stem}_{ts}{db_path.suffix}"
        shutil.copy2(db_path, backup_path)

        # retention
        try:
            backups = sorted([p for p in backups_dir.glob(f"{db_path.stem}_*{db_path.suffix}") if p.is_file()],
                             key=lambda p: p.stat().st_mtime, reverse=True)
            for old in backups[keep_last:]:
                try:
                    old.unlink()
                except Exception:
                    pass
        except Exception:
            pass
    except Exception:
        pass

def _atomic_replace(tmp_path: Path, final_path: Path):
    """Sostituisce final_path con tmp_path in modo atomico (locale)."""
    final_path.parent.mkdir(parents=True, exist_ok=True)
    os.replace(str(tmp_path), str(final_path))

def _meta_path_for(db_path: Path) -> Path:
    return db_path.with_name("db_meta.json")

def _bump_db_version(db_path: Path):
    """Incrementa un contatore versione DB (utile per cache/invalidation)."""
    try:
        mp = _meta_path_for(db_path)
        meta = {}
        if mp.exists():
            try:
                meta = json.loads(mp.read_text(encoding="utf-8"))
            except Exception:
                meta = {}
        meta["db_version"] = int(meta.get("db_version", 0)) + 1
        meta["last_write_ts"] = datetime.now().isoformat(timespec="seconds")
        tmp = mp.with_suffix(".tmp")
        tmp.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(str(tmp), str(mp))
    except Exception:
        pass



def _excel_engine_for_name(name: str | None) -> str | None:
    """Best-effort engine selection for pandas Excel readers.

    Pandas may fail to infer the engine when the input is a file-like object
    (e.g. Streamlit UploadedFile) or when the file has a non-standard extension.
    """
    if not name:
        return None
    suffix = Path(str(name)).suffix.lower()
    if suffix in {'.xlsx', '.xlsm', '.xltx', '.xltm'}:
        return 'openpyxl'
    # .xls requires xlrd (often not installed). Prefer converting to .xlsx.
    if suffix == '.xls':
        return 'xlrd'
    if suffix == '.xlsb':
        return 'pyxlsb'
    if suffix == '.ods':
        return 'odf'
    return None


def _excel_engine_for_obj(obj) -> str | None:
    """Try to determine the engine from a path or Streamlit UploadedFile."""
    try:
        if isinstance(obj, (str, Path)):
            return _excel_engine_for_name(str(obj))
        name = getattr(obj, 'name', None)
        return _excel_engine_for_name(name)
    except Exception:
        return None
import streamlit as st

# ============================
# Import robusto (anche senza intestazioni)
# ============================

def _safe_str(x) -> str:
    if pd.isna(x):
        return ""
    try:
        return str(x).strip()
    except Exception:
        return ""

def _series_date_score(s: pd.Series) -> float:
    """Ritorna una score [0..1] di quanto la serie sembri una data."""
    if s is None or len(s) == 0:
        return 0.0
    dt = pd.to_datetime(s, errors='coerce', dayfirst=True)
    ok = dt.notna().mean()
    if ok == 0:
        return 0.0
    # penalizza date fuori range
    years = dt.dropna().dt.year
    if len(years) == 0:
        return 0.0
    in_range = ((years >= 2000) & (years <= 2100)).mean()
    return float(ok * in_range)

def _series_numeric_score(s: pd.Series) -> float:
    if s is None or len(s) == 0:
        return 0.0
    num = pd.to_numeric(s, errors='coerce')
    ok = num.notna().mean()
    return float(ok)

def _series_name_score(s: pd.Series) -> float:
    """Score per colonna NOME: tante lettere/spazi, pochi numeri, lunghezze medie."""
    if s is None or len(s) == 0:
        return 0.0
    ss = s.astype(str).fillna("")
    sample = ss.head(500)
    if len(sample) == 0:
        return 0.0
    has_space = sample.str.contains(r"\s", regex=True).mean()
    has_digit = sample.str.contains(r"\d", regex=True).mean()
    avg_len = sample.str.len().mean()
    # nomi spesso >= 6 caratteri, con spazi, senza digit
    score = (has_space * 0.6) + ((1 - has_digit) * 0.3) + (min(avg_len / 12.0, 1.0) * 0.1)
    return float(max(0.0, min(1.0, score)))

def _series_matricola_score(s: pd.Series) -> float:
    """Score per colonna MATRICOLA: spesso contiene cifre, spesso senza spazi, abbastanza univoca."""
    if s is None or len(s) == 0:
        return 0.0
    ss = s.astype(str).fillna("")
    sample = ss.head(800)
    if len(sample) == 0:
        return 0.0
    has_digit = sample.str.contains(r"\d", regex=True).mean()
    no_space = (1 - sample.str.contains(r"\s", regex=True).mean())
    # pattern tipico: 3-6 cifre + eventuale lettera (es 3371N) oppure alfanumerico corto
    pat = sample.str.match(r"^(?=.{3,10}$)(?=.*\d)[A-Za-z0-9]+$", na=False).mean()
    uniq = sample.nunique(dropna=True) / max(1, len(sample))
    score = (has_digit * 0.35) + (no_space * 0.25) + (pat * 0.25) + (min(uniq / 0.9, 1.0) * 0.15)
    return float(max(0.0, min(1.0, score)))

def _series_turno_score(s: pd.Series) -> float:
    """Score per colonna TURNO: codici come M78, N11, P38, FER, RPD, MAL..."""
    if s is None or len(s) == 0:
        return 0.0
    ss = s.astype(str).fillna("").str.strip().str.upper()
    sample = ss.head(800)
    if len(sample) == 0:
        return 0.0
    pat1 = sample.str.match(r"^[A-Z]{1,2}\d{1,3}$", na=False).mean()
    known = sample.isin({"FER", "RPD", "MAL", "ASS", "RIP", "RIPO", "PER", "P"}).mean()
    short = (sample.str.len().between(2, 5)).mean()
    score = (pat1 * 0.6) + (known * 0.25) + (short * 0.15)
    return float(max(0.0, min(1.0, score)))

def _looks_like_header_row(df: pd.DataFrame) -> bool:
    """Heuristica: la prima riga contiene parole chiave tipo Data/Matricola/Turno/Ore."""
    if df is None or df.empty:
        return False
    first = df.iloc[0].astype(str).fillna("").str.strip().str.lower()
    joined = " | ".join(first.tolist())
    keys = ["data", "matric", "turn", "ore", "min", "valore", "nome", "cognome"]
    return any(k in joined for k in keys)

def _normalize_attivita_headerless(df: pd.DataFrame) -> pd.DataFrame:
    """Normalizza un DataFrame **Attivita** senza intestazioni (header=None).

    Formato atteso (posizionale, come export GT):
      0 = Cognome Nome (testo)
      1 = Matricola (alfanumerico con cifre)
      2 = Unita Operativa (UO)
      3 = Turno primario (primary shift)
      4 = Attivita secondaria (ATT)
      5 = Data (dd/mm/yyyy)  <-- dayfirst=True
      6 = Valore in minuti (numerico) -> convertito in ore
      7 = POX (posizione operativa) / note operative (testo)

    Output colonne: nome, matricola, uo, turno, att, pox, data, minuti, valore
      - minuti: intero (minuti)
      - valore: ore float (2 decimali)
    """
    if df is None or df.empty:
        return pd.DataFrame(columns=["nome", "matricola", "turno", "data", "valore"])

    # rimuovi colonne completamente vuote
    df = df.dropna(axis=1, how='all').copy()

    # Mapping posizionale (preferito): foglio "attivita" headerless
    fallback = {}
    if df.shape[1] >= 7:
        fallback = {
            "nome": 0,
            "matricola": 1,
            "uo": 2,
            "turno": 3,
            "att": 4,
            "data": 5,
            "minuti": 6,
            "pox": 7 if df.shape[1] >= 8 else None,
        }

    cols = list(df.columns)

    # trova data (dd/mm/yyyy -> dayfirst)
    date_scores = {c: _series_date_score(df[c]) for c in cols}
    date_col = max(date_scores, key=date_scores.get) if date_scores else None
    if date_col is None or date_scores.get(date_col, 0) < 0.6:
        date_col = fallback.get("data")

    # trova minuti (colonna numerica)
    num_scores = {c: _series_numeric_score(df[c]) for c in cols}
    # escludi colonna data se convertibile a numerico (excel date serial)
    if date_col in num_scores:
        num_scores[date_col] *= 0.2
    minuti_col = max(num_scores, key=num_scores.get) if num_scores else None
    if minuti_col is None or num_scores.get(minuti_col, 0) < 0.6:
        minuti_col = fallback.get("minuti")

    # trova matricola (colonna con numeri)
    mat_scores = {c: _series_matricola_score(df[c]) for c in cols}
    for c in [date_col, minuti_col]:
        if c in mat_scores:
            mat_scores[c] *= 0.1
    matricola_col = max(mat_scores, key=mat_scores.get) if mat_scores else None
    if matricola_col is None or mat_scores.get(matricola_col, 0) < 0.45:
        matricola_col = fallback.get("matricola")

    # trova nome
    name_scores = {c: _series_name_score(df[c]) for c in cols}
    for c in [date_col, minuti_col, matricola_col]:
        if c in name_scores:
            name_scores[c] *= 0.1
    nome_col = max(name_scores, key=name_scores.get) if name_scores else None
    if nome_col is None or name_scores.get(nome_col, 0) < 0.35:
        nome_col = fallback.get("nome")

    # trova turno
    turno_scores = {c: _series_turno_score(df[c]) for c in cols}
    for c in [date_col, minuti_col, matricola_col, nome_col]:
        if c in turno_scores:
            turno_scores[c] *= 0.1
    turno_col = max(turno_scores, key=turno_scores.get) if turno_scores else None
    if turno_col is None or turno_scores.get(turno_col, 0) < 0.25:
        turno_col = fallback.get("turno")

    # campi aggiuntivi (UO, ATT, POX) se presenti
    uo_col = fallback.get("uo")
    att_col = fallback.get("att")
    pox_col = fallback.get("pox")

    out = pd.DataFrame({
        "nome": df[nome_col] if nome_col in df.columns else "",
        "matricola": df[matricola_col] if matricola_col in df.columns else "",
        "uo": df[uo_col] if (uo_col in df.columns) else "",
        "turno": df[turno_col] if turno_col in df.columns else "",
        "att": df[att_col] if (att_col in df.columns) else "",
        "pox": df[pox_col] if (pox_col in df.columns) else "",
        "data": df[date_col] if date_col in df.columns else pd.NaT,
        "minuti": df[minuti_col] if minuti_col in df.columns else 0,
    })

    # pulizia
    out["matricola"] = out["matricola"].apply(_safe_str)
    out["nome"] = out["nome"].apply(_safe_str)
    out["turno"] = out["turno"].apply(_safe_str)
    out["uo"] = out["uo"].apply(_safe_str)
    out["att"] = out["att"].apply(_safe_str)
    out["pox"] = out["pox"].apply(_safe_str)
    out["data"] = pd.to_datetime(out["data"], errors='coerce', dayfirst=True)

    # minuti -> ore
    m = pd.to_numeric(out["minuti"], errors='coerce')
    m = m.fillna(0.0)
    out["minuti"] = m.astype(float)
    out["valore"] = (out["minuti"] / 60.0).round(2)

    # rimuovi righe completamente vuote
    out = out[(out["matricola"] != "") | (out["nome"] != "") | (out["turno"] != "") | (out["valore"] != 0) | (out["data"].notna())]

    return out

def _normalize_columns_generic(df: pd.DataFrame) -> pd.DataFrame:
    """Normalizza colonne comuni: case/strip e alias minimi."""
    if df is None:
        return pd.DataFrame()
    df = df.copy()

    def _make_unique_columns(cols):
        """Rende univoci i nomi colonna preservando l'ordine."""
        seen = {}
        out = []
        for c in cols:
            base = str(c).strip()
            if base not in seen:
                seen[base] = 0
                out.append(base)
            else:
                seen[base] += 1
                out.append(f"{base}_{seen[base]}")
        return out

    # prima di qualsiasi rename, rendi univoci (Excel puo' avere header duplicati)
    df.columns = _make_unique_columns(df.columns)
    rename = {}
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in {"data", "date", "giorno"}:
            rename[c] = "data"
        elif "matric" in cl or cl in {"badge", "id"}:
            rename[c] = "matricola"
        elif "turn" in cl or cl in {"shift"}:
            rename[c] = "turno"
        # IMPORTANTE: non collassare 'minuti' dentro 'valore' (crea duplicati)
        elif cl in {"minuti", "minutes"}:
            rename[c] = "minuti"
        elif cl in {"ore", "hours"}:
            rename[c] = "ore"
        elif cl in {"valore", "value"}:
            rename[c] = "valore"
        elif cl in {"nome", "nominativo"}:
            rename[c] = "nome"
    if rename:
        df = df.rename(columns=rename)

    # post-rename: garantisci ancora univocita'
    if df.columns.duplicated().any():
        df.columns = _make_unique_columns(df.columns)
    return df


# ============================
# Template intestazioni tabelle (vuote)
# ============================

TEMPLATE_HEADERS = {
    'AbilitazioniTipo': ['abilitazione', 'Tipologia', 'note', 'periodicita1', 'periodicita2'],
    'AltreAbilitazioniPers': ['Matricola', 'categ_professionale', 'Abilitazione', 'note', 'dataSuCert', 'rinnovo', 'zone', 'nr_documento'],
    'CatProfTipo': ['CAT', 'note'],
    'ColoriTurni': ['Pattern', 'BkR', 'BkG', 'BkB', 'FkR', 'FkG', 'FkB', 'Bold', 'Priority'],
    'Ferie_AP': ['matricola', 'cognome_e_nome', 'gg_ferie_AP', 'hr_ferie_AP', 'gg_RFS_AP', 'hr_RFS_AP', 'anno'],
    'Ferie_spettanti': ['Da anni', 'A anni', 'Giorni_Spett', 'RFS_Spett'],
    'IdoneitaMedichePers': ['Matricola', 'categ_professionale', 'idoneita', 'note', 'dataSuCert', 'rinnovo'],
    'Personale': ['CAT', 'matricola', 'Nome', 'UO', 'regime_orario', 'no_spec', 'note_scheda_pers', 'In_Forza', 'Data_di_Nascita', 'Luogo_di_Nascita', 'Cittadinanza', 'Via_e_numero_civico', 'Città', 'Provincia', 'CAP', 'Cod_Fiscale', 'note', 'Indirizzo_posta_elettronica', 'Telefono_uff', 'Telefono_ab', 'Tel_Cellulare', 'Fax', 'Data_prima_Assunzione', 'Licenza', 'Data_Rilascio_Licenza', 'Allegato', 'ENG_lvl', 'Documenti'],
    'Personale_PartTime': ['cognome_e_nome', 'matricola', 'categ_professionale', '%_part_time', 'ID_part_time', 'dal', 'al'],
    'Ripo': ['MATRICOLA', 'COGNOME', 'NOME', 'SETTORE', 'DATA', 'IMPIANTO', 'INIZIO TURNO', 'FINE TURNO', 'ORAE', 'ORAU', 'PAUSA', 'SPEZZONE', 'SIGLA_ATTIVITA', 'StartDT', 'EndDT', 'WorkDate', 'Seq', 'MinutiPOX'],
    'SpecUO-CompTipo': ['SpecUO/Comp', 'Tipologia', 'note', '1_R', '2_R', 'R_C', 'Su_impianto'],
    'SpecUOPers': ['Matricola', 'categ_professionale', 'SpecUO/Comp', 'note', 'dataSuLic', '1_rinnovo', 'doc_1r', '2_rinnovo', 'doc_2r', 'R_comp', 'doc_R_comp', 'note1R', 'note2R', 'notecomp'],
    'Specializzazioni_tipo': ['SPECIALIZZAZIONE', 'POX'],
    'Straordinario': ['CAT', 'Nome', 'matricola', 'UO', 'turno', 'att', 'data', 'valore', 'STP', 'contattato_il', 'note'],
    'TargetOp': ['Days', 'TargetOpTot'],
    'Turni_tipo': ['Turno', 'OraInizio', 'OraFine', 'Minuti', 'SplitNotte', 'MinutiGiornoCorrente', 'MinutiGiornoPrecedente', 'IsOperativo', 'Categoria'],
    'idoneitaMedicheTipo': ['Idoneita', 'Tipologia', 'note', 'periodicita1', 'periodicita2', 'Is_Attiva'],
    'tbl_UO': ['UO'],
    'Festivi': ['GiornoFestivo', 'Descrizione', 'Localita'],
    # Tabella editabile: elenco codici che rappresentano assenze (es. FER, MAL, 104, ...)
    'Turni_Assenze': ['Turno'],
}

# Default elenco assenze (editabile dall'utente nella tabella Turni_Assenze)
# Nota: nei dati GT ricorrono spesso FER (ferie) e RPD/RPN (riposi). Se non esclusi,
# i festivi verrebbero conteggiati anche quando la persona non ...
DEFAULT_TURNI_ASSENZE = [
    'FER','RPD','RPN',
    '104','ASS','CONP','DONS','MAL','MAT','PDEC','VMSP','ROS','NASF','INF','MALF',
    'MIX','PRLS','SCIO','RCI','ALL','PELE','PEXT','PSIN','ESAU'
]


def _attivita_move_extra_turno_to_att(df: pd.DataFrame, primary_turni: set[str] | None = None) -> pd.DataFrame:
    """Normalizza la tabella Attivita quando le *attività secondarie* sono state inserite come righe extra nel campo TURNO.

    Scenario tipico (export/import):
      - stessa matricola + stesso giorno -> più righe
      - una riga con turno primario (es. M78/P38/N11...) + minuti ~480
      - altre righe con codici attività (es. ESAU/ISTR/UCS...) e minuti diversi, ma messe in colonna TURNO

    Obiettivo:
      - per ogni (matricola, giorno) tenere 1 sola riga con TURNO valorizzato (primario)
      - convertire le altre righe in secondarie usando la colonna ATT (TURNO vuoto)
      - preservare minuti/valore per il calcolo ore delle secondarie
    """
    if df is None or len(df) == 0:
        return df

    out = df.copy()

    # colonne minime
    for c in ["matricola", "turno", "att"]:
        if c not in out.columns:
            out[c] = ""
    if "data" not in out.columns:
        return out

    # data -> giorno (senza orario)
    out["data"] = pd.to_datetime(out["data"], errors="coerce", dayfirst=True)
    out["_day"] = out["data"].dt.normalize()

    # minuti coerenti
    if "minuti" in out.columns:
        out["minuti"] = pd.to_numeric(out["minuti"], errors="coerce").fillna(0.0).astype(float)
    else:
        # fallback: se esiste valore in ore
        if "valore" in out.columns:
            out["valore"] = pd.to_numeric(out["valore"], errors="coerce").fillna(0.0).astype(float)
            out["minuti"] = (out["valore"] * 60.0).round(0)
        else:
            out["minuti"] = 0.0

    out["turno"] = out["turno"].fillna("").astype(str).str.strip()
    out["att"] = out["att"].fillna("").astype(str).str.strip()

    primary_turni = {str(x).strip().upper() for x in (primary_turni or set()) if str(x).strip()}

    def pick_primary(g: pd.DataFrame) -> int:
        # preferisci codici che appartengono a Turni_tipo, altrimenti max minuti
        t = g["turno"].fillna("").astype(str).str.strip().str.upper()
        is_primary_code = t.isin(primary_turni) if primary_turni else pd.Series([False]*len(g), index=g.index)
        if is_primary_code.any():
            gg = g[is_primary_code]
            # tra i primari scegli quello con più minuti
            return int(gg["minuti"].idxmax())
        # fallback: il record con più minuti
        return int(g["minuti"].idxmax())

    new_rows = []
    for (mat, day), g in out.groupby(["matricola", "_day"], dropna=False):
        if len(g) == 1:
            new_rows.append(g)
            continue

        idx_primary = pick_primary(g)
        primary_row = g.loc[[idx_primary]].copy()

        # le altre righe diventano secondarie
        others = g.drop(index=idx_primary).copy()
        # se att è vuoto, usa turno come codice attività
        others["att"] = others["att"].where(others["att"].astype(str).str.strip().ne(""), others["turno"])
        # turno vuoto sulle secondarie per evitare "più turni primari"
        others["turno"] = ""

        # se la riga primaria aveva anche un'attività in colonna att, la lasciamo; se invece l'attività era in turno (caso raro)
        primary_row["att"] = primary_row["att"].astype(str).str.strip()

        new_rows.append(primary_row)
        new_rows.append(others)

    out2 = pd.concat(new_rows, ignore_index=True)

    # pulizia helper
    out2 = out2.drop(columns=["_day"], errors="ignore")

    # ricalcola valore (ore) da minuti, se presente
    out2["minuti"] = pd.to_numeric(out2.get("minuti", 0), errors="coerce").fillna(0.0).astype(float)
    out2["valore"] = (out2["minuti"] / 60.0).round(2)

    # dedup finale
    out2 = out2.drop_duplicates()

    return out2


class PersGestDatabase:
    """Gestisce il file Excel master come database"""

    # Tabelle del database (create automaticamente se mancanti)
    TABLES = [
        'Attivita',
        'Personale',
        'Straordinario',
        'Ripo',
        'tbl_UO',
        'CatProfTipo',
        'Turni_tipo',
        'ColoriTurni',
        'TargetOp',
        'AbilitazioniTipo',
        'AltreAbilitazioniPers',
        'idoneitaMedicheTipo',
        'IdoneitaMedichePers',
        'Specializzazioni_tipo',
        'SpecUO-CompTipo',
        'SpecUOPers',
        'Personale_PartTime',
        'Ferie_spettanti',
        'Ferie_AP',
        'Ferie_Godute',
        'Malattia',
        'Infortunio',
        'Maternita',
        'Permesso',
        'Aspettativa',
        'Congedo',
        'Sciopero',
        'Altro_Assenza',
        'Formazione',
        'Missione',
        'Smart_Working',
        'Note',
        'Festivi',
        'Turni_Assenze',
    ]
    
    def _default_db_path(self) -> Path:
        """Percorso database persistente (non dentro la cartella del progetto)."""
        # Windows: LOCALAPPDATA\PersGestStreamlit\persgest_master.xlsx
        base = Path(os.environ.get("LOCALAPPDATA", "")) if os.environ.get("LOCALAPPDATA") else (Path.home() / ".persgest")
        p = base / "PersGestStreamlit" / "persgest_master.xlsx" if os.environ.get("LOCALAPPDATA") else (base / "persgest_master.xlsx")
        p.parent.mkdir(parents=True, exist_ok=True)
        return p

    def __init__(self, excel_path='data/persgest_master.xlsx'):
        """Inizializza database Excel.

        - Se viene passato un path esplicito (es. cartella condivisa), viene usato quello.
        - Se viene passato il path di default del progetto (data/persgest_master.xlsx),
          il database viene salvato in un percorso persistente (AppData / home) per non perdere i dati
          durante gli aggiornamenti.
        """
        passed = Path(excel_path) if excel_path else Path('data/persgest_master.xlsx')

        # Se e' il path "standard" del progetto (relativo), usa path persistente
        passed_norm = str(passed).replace('\\', '/')
        if (not passed.is_absolute()) and (passed_norm.endswith('data/persgest_master.xlsx') or passed_norm.endswith('persgest_master.xlsx')):
            persistent = self._default_db_path()
            # Migra una volta se esiste il vecchio DB nel progetto e non esiste ancora quello persistente
            try:
                if passed.exists() and not persistent.exists():
                    persistent.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy2(passed, persistent)
            except Exception:
                pass
            self.excel_path = persistent
        else:
            self.excel_path = passed

        self.excel_path.parent.mkdir(parents=True, exist_ok=True)

        # Engine Excel (preferisci openpyxl per xlsx/xlsm)
        self._excel_engine = _excel_engine_for_obj(self.excel_path) or 'openpyxl'
        
        # Cache per performance
        if 'db_cache' not in st.session_state:
            st.session_state.db_cache = {}
        
        # Inizializza file se non esiste
        if not self.excel_path.exists():
            self._create_empty_database()

        # Se il file esiste ma e' vuoto/corrotto (tipico dopo copie/parziali download),
        # pandas non riesce a determinare il formato. In quel caso lo rigeneriamo.
        try:
            if self.excel_path.exists() and self.excel_path.stat().st_size < 4096:
                # Backup best-effort
                try:
                    bak = self.excel_path.with_suffix(self.excel_path.suffix + ".bak")
                    shutil.copy2(self.excel_path, bak)
                except Exception:
                    pass
                self._create_empty_database()
        except Exception:
            pass

        # Aggiunge eventuali nuove tabelle (fogli) senza richiedere re-import
        self._ensure_tables_exist()
    
def _create_empty_database(self):
    """Crea file Excel vuoto con tutti i fogli previsti (vuoti)."""
    with _persgest_write_lock(self.excel_path):
        _backup_excel(self.excel_path)

        tmp_path = self.excel_path.with_suffix(self.excel_path.suffix + ".tmp")
        try:
            with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
                for table in self.TABLES:
                    # Foglio vuoto
                    cols = TEMPLATE_HEADERS.get(table, [])
                    if table == 'Turni_Assenze' and cols:
                        # Default iniziale (modificabile dall'utente)
                        df = pd.DataFrame({cols[0]: DEFAULT_TURNI_ASSENZE})
                    else:
                        df = pd.DataFrame(columns=cols)
                    df.to_excel(writer, sheet_name=table, index=False)

            _atomic_replace(tmp_path, self.excel_path)
            _bump_db_version(self.excel_path)
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass
    def _ensure_tables_exist(self):
        """Assicura che tutte le tabelle (fogli) esistano nel file Excel.

        - Se mancano fogli, li crea vuoti (0 righe) con intestazioni note (se disponibili).
        - Non sovrascrive i fogli gia' presenti.
        """
        try:
            xl = pd.ExcelFile(self.excel_path, engine=_excel_engine_for_obj(self.excel_path) or self._excel_engine)
            existing = set(xl.sheet_names)
        except Exception:
            existing = set()

        missing = [t for t in self.TABLES if t not in existing]
        if not missing:
            return

        all_data = {}
        # carica fogli esistenti
        for s in sorted(existing):
            try:
                all_data[s] = pd.read_excel(self.excel_path, sheet_name=s, engine=_excel_engine_for_obj(self.excel_path) or self._excel_engine)
            except Exception:
                all_data[s] = pd.DataFrame()

        # Se la tabella Turni_Assenze esiste già ma risulta vuota (tipico quando si aggiorna un DB
        # già esistente con una nuova feature), pre-carica l'elenco di default delle assenze.
        # Non sovrascriviamo mai dati esistenti: inseriamo solo se non c'è alcun valore utile.
        if 'Turni_Assenze' in all_data:
            df_ta = all_data.get('Turni_Assenze', pd.DataFrame())
            if df_ta is None or df_ta.empty:
                all_data['Turni_Assenze'] = pd.DataFrame({'Turno': DEFAULT_TURNI_ASSENZE})
            else:
                col = 'Turno' if 'Turno' in df_ta.columns else (df_ta.columns[0] if len(df_ta.columns) else None)
                if col is not None:
                    vals = df_ta[col].astype(str).str.strip()
                    vals = vals[vals.ne('') & vals.ne('nan')]
                    if len(vals) == 0:
                        all_data['Turni_Assenze'] = pd.DataFrame({'Turno': DEFAULT_TURNI_ASSENZE})

        # aggiungi fogli mancanti vuoti
        for t in missing:
            cols = TEMPLATE_HEADERS.get(t, [])
            if t == 'Turni_Assenze' and cols:
                all_data[t] = pd.DataFrame({'Turno': DEFAULT_TURNI_ASSENZE})
            else:
                all_data[t] = pd.DataFrame(columns=cols)

        
with _persgest_write_lock(self.excel_path):
    _backup_excel(self.excel_path)
    tmp_path = self.excel_path.with_suffix(self.excel_path.suffix + ".tmp")
    try:
        with pd.ExcelWriter(tmp_path, engine='openpyxl') as writer:
            for name, df in all_data.items():
                df.to_excel(writer, sheet_name=name, index=False)

        _atomic_replace(tmp_path, self.excel_path)
        _bump_db_version(self.excel_path)
    finally:
        try:
            if tmp_path.exists():
                tmp_path.unlink()
        except Exception:
            pass
        # reset cache
        try:
            self.get_all.clear()
        except Exception:
            pass
        st.cache_data.clear()

    @st.cache_data(ttl=5)
    def get_all(_self, table):
        """Leggi tutti i record da una tabella
        
        Args:
            table: Nome tabella/foglio
            
        Returns:
            DataFrame con i dati
        """
        if table not in _self.TABLES:
            raise ValueError(f"Tabella {table} non esiste")
        
        try:
            eng = _excel_engine_for_obj(_self.excel_path) or getattr(_self, '_excel_engine', None)
            df = pd.read_excel(_self.excel_path, sheet_name=table, engine=eng)
            # Excel puo' contenere intestazioni duplicate: rendi univoci subito
            if df is not None and not df.empty and getattr(df.columns, 'duplicated', None) is not None:
                if df.columns.duplicated().any():
                    seen = {}
                    new_cols = []
                    for c in df.columns:
                        base = str(c).strip()
                        if base not in seen:
                            seen[base] = 0
                            new_cols.append(base)
                        else:
                            seen[base] += 1
                            new_cols.append(f"{base}_{seen[base]}")
                    df.columns = new_cols
            df = _normalize_columns_generic(df)

            # Fix speciale: Attivita spesso importata senza intestazioni (headerless)
            if table == 'Attivita':
                needed = {'data', 'matricola', 'turno'}
                if not needed.issubset(set(df.columns)):
                    # prova a rileggere come header=None e normalizzare
                    try:
                        df0 = pd.read_excel(_self.excel_path, sheet_name=table, header=None, engine=eng)
                        df = _normalize_attivita_headerless(df0)
                    except Exception:
                        pass
                # garantisci colonne base
                for col in ['nome','matricola','uo','turno','att','pox','data','minuti','valore']:
                    if col not in df.columns:
                        df[col] = '' if col not in {'minuti','valore','data'} else (0 if col in {'minuti','valore'} else pd.NaT)

                # Se esistono ancora colonne duplicate tipo valore_1, valore_2 (import vecchi), consolida.
                val_cols = [c for c in df.columns if str(c).startswith('valore')]
                if len(val_cols) > 1:
                    scores = {}
                    for c in val_cols:
                        s = pd.to_numeric(df[c], errors='coerce').fillna(0.0).abs()
                        scores[c] = float((s > 0).mean())
                    keep = max(scores, key=scores.get)
                    df['valore'] = pd.to_numeric(df[keep], errors='coerce').fillna(0.0)
                    drop_cols = [c for c in val_cols if c != keep and c != 'valore']
                    if drop_cols:
                        df = df.drop(columns=drop_cols)

            # Converti date
            date_cols = ['data', 'data_inizio', 'data_fine']
            for col in date_cols:
                if col in df.columns:
                    df[col] = pd.to_datetime(df[col], errors='coerce', dayfirst=True)

            return df
        except Exception as e:
            st.error(f"Errore lettura {table}: {e}")
            return pd.DataFrame()
    
def save_table(self, table, df):
    """Salva DataFrame su foglio Excel SENZA rischiare di svuotare le altre tabelle.

    Strategia: aggiorna solo il foglio richiesto via openpyxl, preservando gli altri fogli.

    **Safety per uso multi-utente (≈10 utenti)**:
    - serializza le scritture con lock (globale + su file)
    - crea backup automatico prima della scrittura
    - salva su file temporaneo e sostituisce in modo atomico (os.replace)
    """
    if table not in self.TABLES:
        raise ValueError(f"Tabella {table} non esiste")

    # Import locali per evitare dipendenze in fase di import modulo
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Normalizza dataframe
    if df is None:
        df = pd.DataFrame()
    try:
        df = df.copy()
    except Exception:
        pass

    # Lock su scrittura: impedisce sovrascritture concorrenti
    with _persgest_write_lock(self.excel_path):
        # Backup best-effort (non altera logiche / strutture)
        _backup_excel(self.excel_path)

        # Carica o crea workbook
        if os.path.exists(self.excel_path):
            wb = load_workbook(self.excel_path)
        else:
            wb = Workbook()
            # rimuovi sheet di default
            try:
                if wb.active and wb.active.title == 'Sheet':
                    wb.remove(wb.active)
            except Exception:
                pass

        # Assicura che TUTTE le tabelle esistano come fogli (senza sovrascriverle)
        for tbl in self.TABLES:
            if tbl not in wb.sheetnames:
                wb.create_sheet(tbl)

        # Rimuovi e ricrea il foglio target alla posizione coerente con TABLES
        try:
            idx = self.TABLES.index(table)
        except Exception:
            idx = 0

        if table in wb.sheetnames:
            ws_old = wb[table]
            wb.remove(ws_old)

        ws = wb.create_sheet(table, index=min(idx, len(wb.sheetnames)))

        # Scrivi header + righe
        for _r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)

        # Salvataggio atomico: tmp -> replace
        tmp_path = self.excel_path.with_suffix(self.excel_path.suffix + ".tmp")
        try:
            wb.save(tmp_path)
            # validazione minima: il file deve riaprirsi
            try:
                _ = load_workbook(tmp_path, read_only=True)
            except Exception as e:
                raise RuntimeError(f"File Excel temporaneo non valido: {e}") from e

            _atomic_replace(tmp_path, self.excel_path)
            _bump_db_version(self.excel_path)
        finally:
            try:
                if tmp_path.exists():
                    tmp_path.unlink()
            except Exception:
                pass

    # Invalida cache
    try:
        self.get_all.clear()
    except Exception:
        pass
    try:
        st.cache_data.clear()
    except Exception:
        pass
    def import_excel(self, uploaded_file, table_mapping, mode: str = 'replace'):
        """Import da file Excel esterno
        
        Args:
            uploaded_file: File Excel caricato
            table_mapping: Dict {foglio_origine: tabella_destinazione}
        """
        try:
            mode = (mode or 'replace').strip().lower()
            if mode not in {'replace', 'append'}:
                mode = 'replace'
            # Streamlit UploadedFile e' un file-like: assicurati che il puntatore sia all'inizio
            try:
                uploaded_file.seek(0)
            except Exception:
                pass
            eng = _excel_engine_for_obj(uploaded_file)
            excel_file = pd.ExcelFile(uploaded_file, engine=eng) if eng else pd.ExcelFile(uploaded_file)

            for sheet_name, dest_table in table_mapping.items():
                if sheet_name not in excel_file.sheet_names:
                    continue

                # --- ATTIVITA (GT_IMPORT) ---
                if dest_table == 'Festivi':
                    # Foglio Festivi: prima colonna può essere numerica tipo 0101 (ggmm). Normalizziamo in gg/mm (anno perpetuo).
                    df_raw = excel_file.parse(sheet_name=sheet_name, header=0)
                    df_raw = _normalize_columns_generic(df_raw)

                    # colonne attese: data (ggmm o date), nome/descrizione (opzionale)
                    cols = {c.lower(): c for c in df_raw.columns}
                    c_data = cols.get("data") or cols.get("giorno") or cols.get("ggmm") or list(df_raw.columns)[0]
                    c_nome = cols.get("nome") or cols.get("festivo") or cols.get("descrizione") or (list(df_raw.columns)[1] if len(df_raw.columns) > 1 else None)

                    def _to_ddmm(v):
                        if v is None or (isinstance(v, float) and pd.isna(v)):
                            return None
                        s = str(v).strip()
                        if not s:
                            return None
                        # Se Excel ha interpretato come data vera
                        try:
                            dt = pd.to_datetime(v, errors="coerce")
                            if pd.notna(dt):
                                return dt.strftime("%d/%m")
                        except Exception:
                            pass
                        s = re.sub(r"[^0-9]", "", s)
                        if len(s) == 3:  # es 101 -> 0101
                            s = "0" + s
                        if len(s) >= 4:
                            gg = s[:2]
                            mm = s[2:4]
                            return f"{gg}/{mm}"
                        return None

                    out = pd.DataFrame()
                    out["ddmm"] = df_raw[c_data].apply(_to_ddmm)
                    out["nome"] = (df_raw[c_nome] if c_nome else "").astype(str).str.strip()
                    out = out[out["ddmm"].notna()].copy()
                    # compatibilità: tabella Festivi usa colonna 'data' come dd/mm
                    out.rename(columns={"ddmm": "data"}, inplace=True)

                    df = out

                elif dest_table == 'Attivita':
                    # Il file "Attivita" può avere header "strano" o mancante: rileviamo se la prima riga è header.
                    preview = excel_file.parse(sheet_name, header=None, nrows=2)
                    if _looks_like_header_row(preview):
                        df = excel_file.parse(sheet_name, header=0)
                        df = _normalize_columns_generic(df)
                    else:
                        df0 = excel_file.parse(sheet_name, header=None)
                        df = _normalize_attivita_headerless(df0)

                else:
                    df = excel_file.parse(sheet_name, header=0)
                    df = _normalize_columns_generic(df)
                # Normalizzazione extra per Attivita:
                # se attività secondarie sono state inserite come righe extra in colonna TURNO,
                # le spostiamo in ATT per evitare "più turni primari" e per mostrarle correttamente nel Crosstab.
                if dest_table == 'Attivita':
                    try:
                        tdf = self.get_all('Turni_tipo')
                        # prova a trovare colonna codice turno
                        tcols = {c.lower(): c for c in (tdf.columns if isinstance(tdf, pd.DataFrame) else [])}
                        c_turno = tcols.get('turno') or tcols.get('codice') or tcols.get('sigla') or None
                        primary_turni = set()
                        if c_turno and len(tdf) > 0:
                            primary_turni = set(tdf[c_turno].dropna().astype(str).str.strip().str.upper().tolist())
                        df = _attivita_move_extra_turno_to_att(df, primary_turni=primary_turni)
                    except Exception:
                        # best effort: non bloccare l'import se qualcosa non torna
                        pass

                df = df.drop_duplicates()

                # Modalita': replace (default) o append
                if mode == 'append':
                    cur = self.get_all(dest_table)
                    if len(cur) > 0:
                        df = pd.concat([cur, df], ignore_index=True)
                        df = df.drop_duplicates()

                        if dest_table == 'Attivita':
                            try:
                                tdf = self.get_all('Turni_tipo')
                                tcols = {c.lower(): c for c in (tdf.columns if isinstance(tdf, pd.DataFrame) else [])}
                                c_turno = tcols.get('turno') or tcols.get('codice') or tcols.get('sigla') or None
                                primary_turni = set()
                                if c_turno and len(tdf) > 0:
                                    primary_turni = set(tdf[c_turno].dropna().astype(str).str.strip().str.upper().tolist())
                                df = _attivita_move_extra_turno_to_att(df, primary_turni=primary_turni)
                            except Exception:
                                pass

                self.save_table(dest_table, df)

            return True, "Import completato!"

        except Exception as e:
            return False, f"Errore import: {e}"
    
    def export_excel(self, tables=None):
        """Export tabelle selezionate in file Excel
        
        Args:
            tables: Lista tabelle da esportare (None = tutte)
            
        Returns:
            Path del file esportato
        """
        if tables is None:
            tables = self.TABLES
        
        export_path = Path(f'data/export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        
        with pd.ExcelWriter(export_path, engine='xlsxwriter') as writer:
            for table in tables:
                df = self.get_all(table)
                df.to_excel(writer, sheet_name=table, index=False)
        
        return export_path
    
    def get_stats(self):
        """Statistiche database
        
        Returns:
            Dict con conteggi per tabella
        """
        stats = {}
        for table in self.TABLES:
            df = self.get_all(table)
            stats[table] = len(df)
        return stats
    
    def clear_table(self, table):
        """Svuota una tabella
        
        Args:
            table: Nome tabella da svuotare
        """
        self.save_table(table, pd.DataFrame())
    
    def add_record(self, table, record):
        """Aggiungi singolo record
        
        Args:
            table: Nome tabella
            record: Dict con dati record
        """
        df = self.get_all(table)
        new_df = pd.concat([df, pd.DataFrame([record])], ignore_index=True)
        self.save_table(table, new_df)
    
    def update_record(self, table, index, record):
        """Aggiorna record esistente
        
        Args:
            table: Nome tabella
            index: Indice riga da aggiornare
            record: Dict con nuovi dati
        """
        df = self.get_all(table)
        for key, value in record.items():
            df.at[index, key] = value
        self.save_table(table, df)
    
    def delete_record(self, table, index):
        """Elimina record
        
        Args:
            table: Nome tabella
            index: Indice riga da eliminare
        """
        df = self.get_all(table)
        df = df.drop(index).reset_index(drop=True)
        self.save_table(table, df)